
from openpyxl import load_workbook, Workbook
from typing import List, Generator
from pathlib import Path
import logging

from config import Applicant

logger = logging.getLogger(__name__)


class DataHandler:
    # Expected column mapping (case-insensitive)
    COLUMN_MAP = {
        "country": ["country", "nationality", "nation"],
        "passport_number": ["passport number", "passport", "passport no", "passport_number"],
        "visa_number": ["visa number", "visa", "visa no", "visa_number"],
        "mobile": ["mobile", "phone", "primary mobile", "mobile number", "contact"],
        "email": ["email", "primary email", "email address", "e-mail"]
    }
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.workbook = None
        self.header_map: dict = {}
    
    def _normalize(self, text: str) -> str:
        """Normalize header text for comparison"""
        return text.lower().strip().replace("_", " ")
    
    def _find_column(self, headers: List[str], field: str) -> int:
        """Find column index for a field"""
        possible_names = self.COLUMN_MAP.get(field, [field])
        for idx, header in enumerate(headers):
            if self._normalize(header) in possible_names:
                return idx
        return -1
    
    
    def _sanitize_phone(self, phone: str) -> str:
        """Sanitize phone to 0092XXXXXXXXXX format (exactly 14 digits)"""
        clean = "".join(filter(str.isdigit, phone))
        
        # If empty, return as is (will be caught by validator)
        if not clean:
            return ""
        
        # Handle various input formats
        if clean.startswith("0092"):
            # Already correct prefix
            pass
        elif clean.startswith("92"):
            # Missing leading 00
            clean = "00" + clean
        elif clean.startswith("0"):
            # Local format (0300...) - convert to international
            clean = "0092" + clean[1:]
        else:
            # Raw number (3001234567) - add full prefix
            clean = "0092" + clean
        
        # Truncate to exactly 14 digits
        return clean[:14]
    
    def _detect_headers(self, sheet) -> dict:
        """Detect column headers and create mapping"""
        first_row = [str(cell.value or "").strip() for cell in sheet[1]]
        
        mapping = {}
        for field in self.COLUMN_MAP.keys():
            col_idx = self._find_column(first_row, field)
            if col_idx >= 0:
                mapping[field] = col_idx
            else:
                logger.warning(f"Column not found for field: {field}")
        
        return mapping
    
    def load(self) -> List[Applicant]:
        """Load all applicants from Excel"""
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")
        
        self.workbook = load_workbook(self.excel_path, read_only=True)
        sheet = self.workbook.active
        
        self.header_map = self._detect_headers(sheet)
        
        # Validate required columns
        required = {"country", "passport_number", "visa_number", "mobile", "email"}
        missing = required - set(self.header_map.keys())
        if missing:
            raise ValueError(f"Missing required columns: {missing}")
        
        applicants = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not any(row):
                continue
            
            try:
                applicant = Applicant(
                    country=str(row[self.header_map["country"]] or "").strip(),
                    passport_number=str(row[self.header_map["passport_number"]] or "").strip(),
                    visa_number=str(row[self.header_map["visa_number"]] or "").strip(),
                    mobile=self._sanitize_phone(str(row[self.header_map["mobile"]] or "")),
                    email=str(row[self.header_map["email"]] or "").strip(),
                    row_index=row_idx
                )
                
                # Validate non-empty
                if all([applicant.passport_number, applicant.visa_number, 
                       applicant.mobile, applicant.email]):
                    applicants.append(applicant)
                else:
                    logger.warning(f"Row {row_idx}: Missing required data, skipping")
                    
            except Exception as e:
                logger.error(f"Row {row_idx}: Failed to parse - {e}")
        
        logger.info(f"Loaded {len(applicants)} valid applicants")
        return applicants
    
    def iterate(self) -> Generator[Applicant, None, None]:
        """Generator for memory-efficient iteration"""
        for applicant in self.load():
            yield applicant
    
    def close(self):
        """Close workbook"""
        if self.workbook:
            self.workbook.close()


class JsonDataHandler:
    """
    Load applicants from JSON file (for web interface).
    Provides the same interface as DataHandler for compatibility.
    """
    
    JSON_FILE = Path(__file__).parent / "applicants.json"
    
    def __init__(self, json_path: str = None):
        self.json_path = Path(json_path) if json_path else self.JSON_FILE
    
    def _sanitize_phone(self, phone: str) -> str:
        """Ensure phone has proper format"""
        clean = "".join(filter(str.isdigit, phone))
        if not clean:
            return ""
        if clean.startswith("00"):
            return clean
        return "00" + clean
    
    def load(self) -> List[Applicant]:
        """Load all pending applicants from JSON"""
        if not self.json_path.exists():
            logger.warning(f"JSON file not found: {self.json_path}")
            return []
        
        try:
            import json
            with open(self.json_path, 'r') as f:
                data = json.load(f)
            
            applicants = []
            for idx, item in enumerate(data.get("applicants", []), start=1):
                # Only load pending applicants
                if item.get("status", "pending") != "pending":
                    continue
                
                try:
                    applicant = Applicant(
                        country=item.get("country", "Pakistan"),
                        passport_number=item.get("passport_number", "").strip(),
                        visa_number=item.get("visa_number", "").strip(),
                        mobile=self._sanitize_phone(item.get("mobile", "")),
                        email=item.get("email", "").strip(),
                        row_index=idx
                    )
                    
                    if all([applicant.passport_number, applicant.visa_number,
                           applicant.mobile, applicant.email]):
                        applicants.append(applicant)
                    else:
                        logger.warning(f"Applicant {idx}: Missing required data, skipping")
                except Exception as e:
                    logger.error(f"Applicant {idx}: Failed to parse - {e}")
            
            logger.info(f"Loaded {len(applicants)} pending applicants from JSON")
            return applicants
            
        except Exception as e:
            logger.error(f"Failed to load JSON: {e}")
            return []
    
    def iterate(self) -> Generator[Applicant, None, None]:
        """Generator for applicants"""
        for applicant in self.load():
            yield applicant
    
    def update_status(self, passport_number: str, status: str, last_booked: str = None):
        """Update applicant status in JSON file"""
        import json
        
        if not self.json_path.exists():
            return False
        
        try:
            with open(self.json_path, 'r') as f:
                data = json.load(f)
            
            for applicant in data.get("applicants", []):
                if applicant.get("passport_number", "").upper() == passport_number.upper():
                    applicant["status"] = status
                    if last_booked:
                        applicant["last_booked"] = last_booked
                    break
            
            with open(self.json_path, 'w') as f:
                json.dump(data, f, indent=2)
            
            logger.info(f"Updated {passport_number} status to {status}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to update status: {e}")
            return False
    
    def close(self):
        """No-op for compatibility"""
        pass


def create_template(output_path: str = "applicants.xlsx"):
    """Create a sample Excel template"""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Applicants"
    
    # Headers
    headers = ["Country", "Passport Number", "Visa Number", "Primary Mobile", "Primary Email"]
    sheet.append(headers)
    
    # Sample data
    sample_data = [
        ["Pakistan", "AB1234567", "QV2024001", "+923001234567", "applicant1@email.com"],
        ["India", "CD9876543", "QV2024002", "+919876543210", "applicant2@email.com"],
        ["Bangladesh", "EF5555555", "QV2024003", "+8801712345678", "applicant3@email.com"],
    ]
    
    for row in sample_data:
        sheet.append(row)
    
    # Auto-adjust column widths
    for col in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        sheet.column_dimensions[col[0].column_letter].width = max_length + 2
    
    wb.save(output_path)
    logger.info(f"Template created: {output_path}")
    return output_path


if __name__ == "__main__":
    # Create sample template
    create_template()
    
    # Test loading
    handler = DataHandler("applicants.xlsx")
    for applicant in handler.iterate():
        print(f"{applicant.passport_number} - {applicant.email}")
    handler.close()
